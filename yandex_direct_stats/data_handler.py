import pandas as pd
import datetime as dt
import re
from string import ascii_uppercase

from tapi_yandex_direct import YandexDirect

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

from typing import List, Tuple, Union, Optional
from decimal import Decimal
from gspread import Worksheet


def report_wrapper(login: str,
                   token: str,
                   fieldnames: List[str] = ("CampaignName", "Impressions", "Clicks", "Ctr", "AvgCpc", "Cost"),
                   date_range_api: str = 'LAST_7_DAYS',
                   date_range_int: int = 7,
                   report_type: str = 'CAMPAIGN_PERFORMANCE_REPORT',
                   report_name: str = 'Performance report',
                   order_by: str = 'CampaignName',
                   client_name: str = '',
                   ) -> None:
    date = date_range_exclude_today(date_range=date_range_int)
    filename = report_filename(date_range=date_range_int, client_name=client_name)
    report_data = get_report(login=login,
                             token=token,
                             fieldnames=fieldnames,
                             date_range=date_range_api,
                             report_type=report_type,
                             report_name=report_name,
                             order_by=order_by)

    report_data_df = report_data_to_dataframe(report_data)
    report_data_df = add_total_row(report_data_df)
    report_data_df = str_to_numbers(report_data_df)
    report_data_df = rename_df_columns(report_data_df)

    report_workbook = dataframe_to_workbook(report_data_df)
    report_workbook = add_report_description(report_workbook,
                                             client_name=client_name,
                                             date=date)
    report_workbook = style_report_workbook(report_workbook)
    save_report_to_xlsx(report_workbook,
                        filename=filename)


def get_report(login: str,
               token: str,
               fieldnames: List[str] = ("CampaignName", "Impressions", "Clicks", "Ctr", "AvgCpc", "Cost"),
               date_range: str = 'CUSTOM_DATE',
               report_type: str = 'CAMPAIGN_PERFORMANCE_REPORT',
               report_name: str = 'Performance report',
               order_by: str = (),
               sort_order: str = "DESCENDING",
               filter_item = (),
               date_from: str = str(dt.date.today() - dt.timedelta(days=7)),
               date_to: str = str(dt.date.today() - dt.timedelta(days=1)),
               goals: List[str] = [],
               attribution_models: List[str] = []) -> List[List[str]]:
    """
    Получить данные через API запрос в Яндекс Директ
    :param sort_order: порядок сортировки ASCENDING - по возрастанию / DESCENDING - по убыванию
    :param login: логин клиента для агентского аккаунта
    :param token: токен аккаунта
    :param fieldnames: поля отчета https://yandex.ru/dev/direct/doc/reports/fields.html/
    :param date_range: период для отчета https://yandex.ru/dev/direct/doc/reports/period.html/
    :param report_type: тип отчета https://yandex.ru/dev/direct/doc/reports/type.html/
    :param report_name: название отчета
    :param order_by: имена полей (столбцов), по которым требуется отсортировать строки в отчете.
    :param filter_item: фильтрация https://yandex.ru/dev/direct/doc/reports/filters.html
    :param goals: идентификаторы целей https://yandex.ru/support/metrica/general/goals.html
    :param attribution_models: модель атрибуции LC/FC/LSC/LYDC https://yandex.ru/support/direct/statistics/attribution-model.html
    :param date_from: начальная дата отчетного периода YYYY-MM-DD
    :param date_to: конечная дата отчетного периода YYYY-MM-DD
    :return:
    """
    api = YandexDirect(
        # Токен доступа.
        access_token=token,
        # Не будет повторять запрос, если закончаться баллы.
        retry_if_not_enough_units=False,
        # Сделает несколько запросов, если кол-во идентификаторов
        # в условиях фильтрации SelectionCriteria будет больше,
        # чем можно запросить в одном запросе. Работает для метода "get".
        auto_request_generation=True,
        # Если в запросе не будут получены все объекты, сделает еще запросы.
        receive_all_objects=True,
        # Включить песочницу.
        is_sandbox=False,
        # Если вы делаете запросы из под агенсткого аккаунта,
        # вам нужно указать логин аккаунта для которого будете делать запросы.
        login=login
    )
    if date_range == 'CUSTOM_DATE':
        report_data = api.reports().get(
            data={
                "params": {
                    "SelectionCriteria": {
                        "Filter": filter_item,
                        "DateFrom": date_from,
                        "DateTo": date_to
                    },
                    "FieldNames": fieldnames,
                    "OrderBy": [{
                        "Field": order_by,
                        "SortOrder": sort_order
                    }],
                    "Goals": goals,
                    "ReportName": report_name,
                    "ReportType": report_type,
                    "DateRangeType": 'CUSTOM_DATE',
                    "AttributionModels": attribution_models,
                    "Format": "TSV",
                    "IncludeVAT": "YES",
                    "IncludeDiscount": "YES",
                }
            }
        )
    else:
        report_data = api.reports().get(
            data={
                "params": {
                    "SelectionCriteria": {
                        "Filter": filter_item
                    },
                    "FieldNames": fieldnames,
                    "OrderBy": [{
                        "Field": order_by,
                        "SortOrder": sort_order
                    }],
                    "Goals": goals,
                    "ReportName": report_name,
                    "ReportType": report_type,
                    "DateRangeType": date_range,
                    "AttributionModels": attribution_models,
                    "Format": "TSV",
                    "IncludeVAT": "YES",
                    "IncludeDiscount": "YES",
                }
            }
        )
    return report_data().transform()


def report_data_to_dataframe(report_data: List[List[str]]) -> pd.DataFrame:
    return pd.DataFrame(report_data[1:], columns=report_data[0])


def workbook_to_dataframe(report_workbook: Workbook) -> pd.DataFrame:
    return pd.DataFrame(report_workbook.active.values)


def dataframe_to_workbook(df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return wb


def report_data_to_workbook(report_data: List[List[str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    for row in report_data:
        ws.append(row)
    return wb


def style_report_workbook(report_workbook: Workbook,
                          default_width: int = 10,
                          autowidth: bool = False) -> Workbook:
    """
    Применяет стили к Workbook объекту
    """
    report_workbook = font_styling_workbook(report_workbook)
    report_workbook = set_columns_widths(report_workbook=report_workbook,
                                         autowidth=autowidth,
                                         default_width=default_width)
    return report_workbook


def save_report_to_xlsx(report_workbook: Workbook,
                        filename: str) -> None:
    report_workbook.save(filename)


def get_data_from_workbook(report_workbook: Workbook) -> List[Tuple[str]]:
    workbook_data = []
    ws = report_workbook.active
    for row in ws.values:
        workbook_data.append(row)
    return workbook_data


def set_columns_widths(report_workbook: Workbook,
                       autowidth: bool,
                       default_width: int) -> Workbook:
    """
    Установить ширину столбцов в объекте Workbook
    :param report_workbook: openpyxl Workbook object
    :param autowidth: автоматическая регулировка ширины столбцов (по макс. длине ячейки в столбце)
    :param default_width: стандартная ширина столбцов
    """
    report_data = get_data_from_workbook(report_workbook)

    max_width = dict(
        zip(
            ascii_uppercase[:len(report_data[0])],
            [default_width for i in range(len(report_data[0]))]
        )
    )
    max_width['A'] = default_width * 3

    if autowidth:
        for col in range(len(report_data[0])):
            for row in range(len(report_data)):
                width = len(str(report_data[row][col]))
                max_width_loc = max_width[list(sorted(max_width.keys()))[col]]
                if width > int(max_width_loc):
                    max_width[list(sorted(max_width.keys()))[col]] = width

    for i in list(max_width.keys()):
        report_workbook.active.column_dimensions[i].width = max_width[i]

    return report_workbook


def font_styling_workbook(report_workbook: Workbook) -> Workbook:
    """
    Настройка шрифта
    """
    ws = report_workbook.active
    report_data = get_data_from_workbook(report_workbook)
    bold_font = Font(name='Calibri',
                     size=9,
                     bold=True)
    standard_font = Font(name='Calibri',
                         size=9,
                         bold=False)

    for col in ascii_uppercase[:len(report_data[0])]:
        for row in range(1, len(report_data) + 2):
            ws[f'{col}{row}'].font = standard_font

    # Описание
    for cell in ws['1']:
        cell.font = bold_font

    # Заголовок
    for cell in ws['3']:
        cell.font = bold_font

    # Строка ИТОГО
    for cell in ws[f'{len(list(ws)) - 1}']:
        cell.font = bold_font

    return report_workbook


def compute_total_row_from_df_report(report_data_df: pd.DataFrame) -> dict[str, Union[int, float]]:
    total_dict = {}
    for headline in report_data_df.columns:
        total_dict[headline] = ''
    try:
        total_dict['CampaignName'] = ['ИТОГО']
    except NameError:
        pass

    try:
        total_dict['Impressions'] = [sum(list(map(int, report_data_df['Impressions'])))]
    except NameError:
        pass

    try:
        total_dict['Clicks'] = [sum(list(map(int, report_data_df['Clicks'])))]
    except NameError:
        pass

    try:
        total_dict['Cost'] = [sum(list(map(float, report_data_df['Cost'])))]
    except NameError:
        pass

    try:
        total_dict['Ctr'] = [f"{total_dict['Clicks'][0] * 100 / total_dict['Impressions'][0]:.2f}"]
    except NameError:
        pass
    except ZeroDivisionError:
        total_dict['Ctr'] = '0'

    try:
        total_dict['AvgCpc'] = [f"{total_dict['Cost'][0] / total_dict['Clicks'][0]:.2f}"]
    except NameError:
        pass
    except ZeroDivisionError:
        total_dict['AvgCpc'] = '0'

    try:
        conversions_not_null = list(filter(lambda x: x != '--', report_data_df['Conversions']))
    except KeyError:
        pass

    try:
        total_dict['Conversions'] = [sum(list(map(int, conversions_not_null)))]
    except NameError:
        pass

    try:
        total_dict['ConversionRate'] = [f"{total_dict['Conversions'][0] * 100 / total_dict['Clicks'][0]:.2f}"]
    except NameError:
        pass
    except ZeroDivisionError:
        total_dict['ConversionRate'] = '0'

    try:
        total_dict['CostPerConversion'] = [f"{total_dict['Cost'][0] / total_dict['Conversions'][0]:.2f}"]
    except NameError:
        pass
    except ZeroDivisionError:
        total_dict['CostPerConversion'] = '0'

    return total_dict


def add_report_description(report_workbook: Workbook,
                           client_name: str = '',
                           date: str = '') -> Workbook:
    """
    Добавить строку описания отчета
    """
    description = f'Отчет Яндекс Директ для клиента {client_name} {date}'

    ws = report_workbook.active
    ws.insert_rows(idx=1, amount=2)
    ws['A1'] = description
    ws['A1'].font = Font(bold=True)

    return report_workbook


def str_to_numbers(report_data_df: pd.DataFrame) -> pd.DataFrame:
    for col in ['Impressions', 'Clicks', 'Conversions']:
        try:
            report_data_df = report_data_df.replace('--', 0)
            report_data_df[col] = report_data_df[col].astype(int)
            report_data_df = report_data_df.replace(0, '--')
        except KeyError:
            print(f'Столбец {col} не используется')

    for col in ['Ctr', 'AvgCpc', 'ConversionRate', 'CostPerConversion', 'Cost']:
        try:
            report_data_df = report_data_df.replace('--', 0)
            report_data_df[col] = report_data_df[col].astype(float)
            report_data_df = report_data_df.replace(0, '--')
        except KeyError:
            print(f'Столбец {col} не используется')

    return report_data_df


def date_range_exclude_today(date_range: Union[int, str]) -> str:

    if date_range == 'LAST_3_DAYS':
        date_range = 3
    elif date_range == 'LAST_7_DAYS':
        date_range = 7
    elif date_range == 'LAST_MONTH':
        current_date = dt.date.today().year, dt.date.today().month, dt.date.today().day
        last_month_last_day = dt.date(current_date[0], current_date[1], 1) - dt.timedelta(days=1)
        last_month_first_day = dt.date(last_month_last_day.year, last_month_last_day.month, 1)
        last_month_first_day_formatted = last_month_first_day.strftime('%d.%m.%Y')
        last_month_last_day_formatted = last_month_last_day.strftime('%d.%m.%Y')
        return f'{last_month_first_day_formatted} - {last_month_last_day_formatted}'

    yesterday = dt.date.today() - dt.timedelta(days=1)
    date_ago = yesterday - dt.timedelta(days=date_range - 1)

    yesterday_formatted = yesterday.strftime('%d.%m.%Y')
    date_ago_formatted = date_ago.strftime('%d.%m.%Y')

    return f'{date_ago_formatted} – {yesterday_formatted}'


def report_filename(date_range: int, client_name: str = '') -> str:
    date = date_range_exclude_today(date_range)
    return f'Яндекс Директ {client_name} {date}.xlsx'


def rename_df_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={'CampaignName': 'Название кампании',
                              'Impressions': 'Показы',
                              'Clicks': 'Клики',
                              'Ctr': 'CTR (%)',
                              'AvgCpc': 'Ср. цена клика (руб.)',
                              'Conversions': 'Конверсии',
                              'ConversionRate': 'Конверсия (%)',
                              'CostPerConversion': 'Цена цели (руб.)',
                              'Cost': 'Стоимость (руб.)'})


def add_report_date_to_google_sheet(worksheet: Worksheet,
                                    days: int = 1,
                                    start_column: str = 'A',
                                    end_column: str = 'I'):
    """
    Добавить в таблицу Google Sheets период отчета и отформатировать
    :param worksheet: Лист Google Sheets
    :param days: кол-во дней в периоде
    :param start_column: стартовая колонка для объединения ячеек
    :param end_column: конечная колонка для объединения ячеек
    :return: None
    """
    report_date_range = [date_range_exclude_today(days)]
    worksheet.append_row(report_date_range)
    date_range_cell = worksheet.findall(report_date_range[0])[-1].row
    worksheet.merge_cells(f'{start_column}{date_range_cell}:{end_column}{date_range_cell}')
    worksheet.format(f'{start_column}{date_range_cell}',
                     {'horizontalAlignment': 'CENTER',
                      'textFormat': {'bold': True,
                                     'fontSize': 14}
                      }
                     )
    return None


def add_report_headline_to_google_sheets(worksheet: Worksheet,
                                         columns: List[str]):
    """
    Добавить заголовок таблицы отчета, сделать жирным и отценрировать в ячейках
    :param worksheet: лист Google Sheets
    :param columns: названия столбцов
    :return: None
    """
    worksheet.append_row(columns)
    report_headline_row_index = worksheet.findall(columns[0])[-1].row
    worksheet.format(f'A{report_headline_row_index}',
                     {
                         'textFormat': {'bold': True}
                     })
    worksheet.format(f'B{report_headline_row_index}:I{report_headline_row_index}',
                     {
                         'textFormat': {'bold': True},
                         'horizontalAlignment': 'CENTER'
                     })
    return None


def values_for_total_row(report_data: List[List[str]]) -> dict[str, Union[int, Decimal]]:
    """
    Добавить итоговую строку, выделяет жирным, отрцентрировать в строке.
    Есть столбцы, зависящие от существования других столбцов.
    :param report_data: данные отчета
    :param worksheet: лист Google Sheets
    :return: None
    """
    headline, report_body = report_data[0], report_data[1:]

    df = pd.DataFrame(report_body, columns=headline)
    df.replace('--', '', inplace=True)
    summary = dict(compute_total_row_from_df_report(df))

    return summary


def format_last_added_report_in_google_sheets(worksheet: Worksheet, prop: str):
    """
    Форматирует ячейки отчета в Google Sheets
    :param worksheet: лист Google Sheets
    :param prop: опорное значение из первой строки отчета,
                 рекомендуется одно из значений в заголовка
    :return:
    """
    start_row_index = worksheet.findall(prop)[-1].row
    end_row_index = worksheet.findall(worksheet.get_all_values()[-1][-1])[-1].row
    worksheet.format(f'B{start_row_index}:I{end_row_index}',
                     {
                         'horizontalAlignment': 'CENTER'
                     }
                     )
    worksheet.format(f'A{start_row_index}:A{end_row_index}',
                     {
                         'horizontalAlignment': 'LEFT'
                     }
                     )
    return None


def format_summary_row_in_google_sheets(worksheet: Worksheet,
                                        total_row: List[str],
                                        start_column: str = 'A',
                                        end_column: str = 'I') -> None:
    summary_row_index = worksheet.findall(total_row[0])[-1].row
    worksheet.format(f'{start_column}{summary_row_index}:{start_column}{summary_row_index}',
                     {
                         'horizontalAlignment': 'RIGHT',
                         'textFormat': {'bold': True}
                     }
                     )
    worksheet.format(
        f'{ascii_uppercase[ascii_uppercase.index(start_column) + 1]}{summary_row_index}:{end_column}{summary_row_index}',
        {
            'horizontalAlignment': 'CENTER',
            'textFormat': {'bold': True}
        }
        )
    return None
